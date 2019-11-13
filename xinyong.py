#!/usr/bin/env python
# -*- encoding: utf-8 -*-

# 批量查询企业的工商信息,注册号.组织机构代码等

import requests
import ssl
import time
import re
import sys
import urllib
from openpyxl import Workbook, load_workbook


def get_info(get_url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
        'Host': 'xin.baidu.com'
    }
    cookies = {
        'BDUSS': 'JPUmFMdnhDfkwwTG52dXk1dUVzVWZRcjlGVmdnQno5ckRwa2x6c3Jtc3NZcEJkSUFBQUFBJCQAAAAAAAAAAAEAAAAB-b5YssvE8bmkvt~N-AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACzVaF0s1WhdRH'
    }
    res_obj = requests.get(url=get_url, headers=headers,
                           cookies=cookies)  # 创建网页访问对象
    pid_z = re.compile(r'pid=(.+?)"')  # 定义pid正则
    q_name_z = re.compile(r'titleName":"(.+?)"')  # 定义企业名称正则
    pid_list = pid_z.findall(res_obj.text.encode(
        'utf-8').decode('unicode_escape'))  # 匹配pid,返回列表
    q_name_list = q_name_z.findall(res_obj.text.encode(
        'utf-8').decode('unicode_escape'))  # 匹配企业名称,返回列表
    return pid_list, q_name_list


def get_info1(get_url1, q_name, name_word2):
    headers1 = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1; PAR-AL00 Build/HUAWEIPAR-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.132 MQQBrowser/6.2 TBS/044304 Mobile Safari/537.36 MicroMessenger/6.7.3.1360(0x26070333) NetType/WIFI Language/zh_CN Process/tools',
        'Host': 'xin.baidu.com',
        'Referer': get_url
    }
    if pid[0] == 'None':  # 如果搜索不到企业则退出函数
        sheet.cell(row_num, 1, row_num-1)  # 插入ID
        sheet.cell(row_num, 2, name_word2)  # 插入企业名称
        wb.save(sys.path[0] + '/001.xlsx')  # 保存为excel格式
        return

    res_obj1 = requests.get(url=get_url1, headers=headers1)
    res_obj1_text = re.sub(r"[ \f\n\r\t\v]", r"", res_obj1.text, )  # 清理空格,换行符
    qy_info_all = re.finditer(r'content">(.*?)<', res_obj1_text)  # 匹配企业信息
    num = 2     # 定义表格初始插入列位置
    sheet.cell(row_num, 1, row_num-1)  # 插入ID
    sheet.cell(row_num, 2, name_word2)  # 插入企业名称
    name_word2 = name_word2.replace(' ', '')     # 删除常规空格
    name_word2 = name_word2.replace('　', '')   # 删除全角空格
    name_word2 = name_word2.replace(u"（", "(")
    name_word2 = name_word2.replace(u"）", ")")
    print(name_word2)
    if q_name != name_word2:  # 给定的企业和爬取的企业名称不一至时,写入搜索的名称
        sheet.cell(row_num, 3, q_name)  # 给第1行第1列的单元格赋值
    for qy_info in qy_info_all:     # 循环插入企业所有信息
        num += 1
        res_obj2_text = re.sub(r'content">|<', r"",
                               qy_info.group(), )  # 清理企业信息首尾多余字符
        print(row_num,  res_obj2_text)
        sheet.cell(row_num, num+1, res_obj2_text)  # 插入企业信息
    sheet.cell(row_num, 21, get_url1)  # 最后一列插入查询网址
    wb.save(sys.path[0] + '/001.xlsx')  # 保存为excel格式


def excel_info():
    # 打开文件：
    excel = load_workbook(sys.path[0] + '/1.xlsx')
    # 获取sheet：
    global rows, table, sheet, wb, get_url, row_num, name_word2     # 设置为全局变量,供其它函数使用
    table = excel.get_sheet_by_name('Sheet1')  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    # cols = table.max_column  # 获取列数
    excel_hed = [('ID', '公司名', '搜索到的公司名', '注册号', '组织机构代码', '税务登记证号', '法定代表人', '经营状态', '成立日期', '营业期限',
                  '审核/年检日期', '注册资本', '企业类型', '机构类型', '所属行业', '行政区划', '电话号码', '登记机关', '注册地址', '经营范围', '查询网址')]  # 定义表头
    # ------------openpyxl操作excel部分------------------------------
    wb = Workbook()  # 创建文件对象
    sheet = wb.active  # 获取默认sheet
    sheet.title = "工商企业信息"  # 修改sheet名字
    row_num = 1  # 定义初始excel行数
    for row1 in excel_hed:      # 写入表头
        sheet.append(row1)


ssl._create_default_https_context = ssl._create_unverified_context  # 忽略ssl错误
excel_info()    # excel 读取创建


for row in range(1, rows):  # excel按列插入信息
    row_num += 1
    name_word2 = table.cell(row=row+1, column=1).value  # 插入读取的企业名称
    name_word = urllib.parse.quote(name_word2)  # url中文转码
    get_url = 'https://xin.baidu.com/s/l?q=' + name_word    # 构建搜索地址
    # 运行搜索代码并接收返回的pid和搜索到的企业名称(用于跟读取到企业名称对比,是否搜索到不一样的企业)
    pid, q_name = get_info(get_url)
    if pid == []:       # 如果搜索不到此企业,则定义pid为空
        pid = ['None']
        q_name = ['None']
    get_url1 = 'https://xin.baidu.com/m/basic?pid=' + \
        pid[0]      # 定义详情页地址
    print(get_url1)
    get_info1(get_url1, q_name[0], name_word2)       # 利用PID获取详情页
    time.sleep(3)       # 休眠3秒,开始下个循环

